import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import pdfplumber
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from xml.sax.saxutils import escape

def extract_data_from_pdf(pdf_path):
    data = []
    headers = None
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                for i, row in enumerate(table):
                    row = [str(cell).replace('\n', ' ').strip() if cell else '' for cell in row]
                    
                    if not headers and any(row):
                        headers = row
                    elif headers and any(row) and row != headers:
                        if len(row) < len(headers):
                            row.extend([''] * (len(headers) - len(row)))
                        elif len(row) > len(headers):
                            row = row[:len(headers)]
                        data.append(row)
                        
    if headers and data:
        return pd.DataFrame(data, columns=headers)
    return pd.DataFrame()

def process_data(file1_path, file2_path):
    try:
        # Load PDF 1 (Alunos)
        df_alunos = extract_data_from_pdf(file1_path)
        if df_alunos.empty:
            raise ValueError("Não foi possível extrair dados da Lista de Alunos.")
            
        # Load PDF 2 (Frequência)
        df_freq = extract_data_from_pdf(file2_path)
        if df_freq.empty:
            raise ValueError("Não foi possível extrair dados do Relatório de Frequência.")

        # Clean column names
        df_alunos.columns = df_alunos.columns.str.replace('\n', ' ').str.strip()
        df_freq.columns = df_freq.columns.str.replace('\n', ' ').str.strip()

        # Find correct columns dynamically, handling possible slight name variations
        col_nome_alunos = next((col for col in df_alunos.columns if 'nome' in col.lower() and 'responsável' not in col.lower()), df_alunos.columns[1] if len(df_alunos.columns) > 1 else None)
        col_aluno_freq = next((col for col in df_freq.columns if 'aluno' in col.lower()), df_freq.columns[2] if len(df_freq.columns) > 2 else None)

        if not col_nome_alunos or not col_aluno_freq:
            raise ValueError("Não foi possível identificar as colunas de nome dos alunos nas tabelas.")

        # Normalize names for merging (removing spaces and making uppercase to match correctly)
        df_alunos['__merge_key'] = df_alunos[col_nome_alunos].astype(str).str.strip().str.upper()
        df_freq['__merge_key'] = df_freq[col_aluno_freq].astype(str).str.strip().str.upper()

        # Merge the two tables based on student names
        merged_df = pd.merge(df_alunos, df_freq, on='__merge_key', how='inner')
        
        if merged_df.empty:
            raise ValueError("Nenhum aluno em comum foi encontrado entre os dois arquivos. Verifique se os nomes coincidem.")

        # Identify requested columns dynamically
        def find_col(df, keyword):
            return next((col for col in df.columns if keyword.lower() in col.lower()), None)

        col_data = find_col(df_freq, 'data')
        if not col_data: col_data = find_col(merged_df, 'data')
        col_turma = find_col(merged_df, 'turma')
        col_presenca = find_col(merged_df, 'presença') or find_col(merged_df, 'presenca')
        col_resp = find_col(df_alunos, 'responsável') or find_col(df_alunos, 'responsavel')
        if not col_resp: col_resp = find_col(merged_df, 'responsável')
        col_tel = find_col(merged_df, 'telefone') or find_col(merged_df, 'contato')
        col_escola = find_col(merged_df, 'escola')

        # Dictionary to store found columns and their desired new names
        final_cols = {}
        if col_data: final_cols[col_data] = 'Data'
        if col_turma: final_cols[col_turma] = 'Turma'
        final_cols[col_nome_alunos] = 'Nome do Aluno'
        if col_presenca: final_cols[col_presenca] = 'Presença'
        if col_resp: final_cols[col_resp] = 'Nome do Responsável'
        if col_tel: final_cols[col_tel] = 'Contato'
        if col_escola: final_cols[col_escola] = 'Escola'

        # Build final DataFrame and apply names
        final_df = merged_df[list(final_cols.keys())].rename(columns=final_cols)
        
        # Remove empty rows
        final_df = final_df.dropna(how='all')
        
        return final_df

    except Exception as e:
        raise Exception(f"Erro ao processar dados: {str(e)}")

def export_to_pdf(df, output_path):
    # Calculate margins to maximize space
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    
    style_cell = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.black
    )
    
    style_header = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.whitesmoke,
        fontName='Helvetica-Bold'
    )
    
    # Prepare data for ReportLab using Paragraph for automatic text wrapping
    data = []
    
    # Headers
    headers = [Paragraph(escape(str(col)), style_header) for col in df.columns]
    data.append(headers)
    
    # Rows
    for row in df.values:
        row_data = [Paragraph(escape(str(item)) if pd.notna(item) else "", style_cell) for item in row]
        data.append(row_data)

    # Calculate column widths dynamically to fit the page horizontally
    width, height = landscape(A4)
    usable_width = width - 40  # Subtracting margins
    
    col_max_lengths = []
    for col in df.columns:
        # Approximate max length by looking at character counts
        max_len = max([len(str(item)) for item in df[col].astype(str)] + [len(str(col))])
        col_max_lengths.append(max(max_len, 5))
        
    total_len = sum(col_max_lengths)
    col_widths = [(l / total_len) * usable_width for l in col_max_lengths]

    t = Table(data, repeatRows=1, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F0F8FF")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(t)
    doc.build(elements)

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Mesclador de Dados PDF")
        self.root.geometry("550x380")
        self.root.resizable(False, False)
        
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.processed_df = None
        
        self.create_widgets()

    def create_widgets(self):
        # Header
        header = ttk.Label(self.root, text="Ferramenta de Cruzamento de Dados (PDF)", font=("Helvetica", 14, "bold"))
        header.pack(pady=(15, 5))
        
        desc = ttk.Label(self.root, text="Selecione os dois PDFs para extrair e mesclar as tabelas.")
        desc.pack(pady=(0, 15))

        # Frame 1
        frame1 = ttk.LabelFrame(self.root, text="1. Lista de Alunos (PDF)", padding=(10, 10))
        frame1.pack(fill="x", padx=20, pady=5)
        ttk.Entry(frame1, textvariable=self.file1_path, state="readonly").pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(frame1, text="Procurar...", command=lambda: self.browse_file(self.file1_path)).pack(side="right")

        # Frame 2
        frame2 = ttk.LabelFrame(self.root, text="2. Relatório de Frequência (PDF)", padding=(10, 10))
        frame2.pack(fill="x", padx=20, pady=5)
        ttk.Entry(frame2, textvariable=self.file2_path, state="readonly").pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(frame2, text="Procurar...", command=lambda: self.browse_file(self.file2_path)).pack(side="right")

        # Process Button
        self.process_btn = ttk.Button(self.root, text="Cruzar Dados", command=self.process)
        self.process_btn.pack(pady=15)
        
        # Export Buttons
        export_frame = ttk.Frame(self.root)
        export_frame.pack(fill="x", padx=20, pady=5)
        
        self.export_xlsx_btn = ttk.Button(export_frame, text="Exportar XLSX", state="disabled", command=self.export_xlsx)
        self.export_xlsx_btn.pack(side="left", expand=True, padx=5)
        
        self.export_pdf_btn = ttk.Button(export_frame, text="Exportar PDF", state="disabled", command=self.export_pdf_action)
        self.export_pdf_btn.pack(side="left", expand=True, padx=5)

    def browse_file(self, var):
        filename = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
        if filename:
            var.set(filename)

    def process(self):
        f1 = self.file1_path.get()
        f2 = self.file2_path.get()
        
        if not f1 or not f2:
            messagebox.showwarning("Aviso", "Por favor, selecione os dois arquivos PDF antes de prosseguir.")
            return
            
        self.process_btn.config(text="Processando...", state="disabled")
        self.root.update()
        
        try:
            self.processed_df = process_data(f1, f2)
            messagebox.showinfo("Sucesso", f"Dados cruzados com sucesso!\nForam encontrados {len(self.processed_df)} registros em comum.")
            self.export_xlsx_btn.config(state="normal")
            self.export_pdf_btn.config(state="normal")
        except Exception as e:
            messagebox.showerror("Erro de Processamento", str(e))
            self.export_xlsx_btn.config(state="disabled")
            self.export_pdf_btn.config(state="disabled")
        finally:
            self.process_btn.config(text="Cruzar Dados", state="normal")

    def export_xlsx(self):
        if self.processed_df is None: return
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if filepath:
            try:
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    self.processed_df.to_excel(writer, index=False, sheet_name='Frequência')
                    worksheet = writer.sheets['Frequência']
                    
                    max_row = len(self.processed_df) + 1
                    max_col = len(self.processed_df.columns)
                    
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    from openpyxl.utils import get_column_letter

                    # Define the table formatting
                    ref = f"A1:{get_column_letter(max_col)}{max_row}"
                    tab = Table(displayName="TabelaFrequencia", ref=ref)
                    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                    tab.tableStyleInfo = style
                    worksheet.add_table(tab)
                    
                    # Auto-fit columns based on text length
                    for col_num, col in enumerate(self.processed_df.columns, 1):
                        column_letter = get_column_letter(col_num)
                        max_length = max(self.processed_df[col].astype(str).map(len).max(), len(col))
                        worksheet.column_dimensions[column_letter].width = max_length + 2

                messagebox.showinfo("Sucesso", "Arquivo XLSX salvo com formatação de tabela com sucesso!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar arquivo XLSX:\n{e}")

    def export_pdf_action(self):
        if self.processed_df is None: return
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if filepath:
            try:
                export_to_pdf(self.processed_df, filepath)
                messagebox.showinfo("Sucesso", "Arquivo PDF salvo com sucesso!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    style = ttk.Style()
    available_themes = style.theme_names()
    if 'vista' in available_themes:
        style.theme_use('vista')
    elif 'clam' in available_themes:
        style.theme_use('clam')
        
    app = App(root)
    root.mainloop()

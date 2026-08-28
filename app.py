import streamlit as st
import pandas as pd
import pdfplumber
import io
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from xml.sax.saxutils import escape

st.set_page_config(page_title="Mesclador de PDFs", page_icon="📄", layout="centered")

# Hide streamlit menu and footer for cleaner look
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

# Helper functions
def extract_data_from_pdf(pdf_file):
    data = []
    headers = None
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                for i, row in enumerate(table):
                    row = [str(cell).replace('\n', ' ').strip() if cell else '' for cell in row]
                    
                    if not headers and any(row):
                        headers = row
                    elif headers and any(row) and row != headers:
                        if len(row) < len(headers):
                            row.extend([''] * (len(headers) - len(row)))
                        elif len(row) > len(headers):
                            row = row[:len(headers)]
                        data.append(row)
                        
    if headers and data:
        return pd.DataFrame(data, columns=headers)
    return pd.DataFrame()

def process_data(file1, file2):
    try:
        # Load PDF 1 (Alunos)
        df_alunos = extract_data_from_pdf(file1)
        if df_alunos.empty:
            raise ValueError("Não foi possível extrair dados da Lista de Alunos.")
            
        # Load PDF 2 (Frequência)
        df_freq = extract_data_from_pdf(file2)
        if df_freq.empty:
            raise ValueError("Não foi possível extrair dados do Relatório de Frequência.")

        # Clean column names
        df_alunos.columns = df_alunos.columns.str.replace('\n', ' ').str.strip()
        df_freq.columns = df_freq.columns.str.replace('\n', ' ').str.strip()

        # Find correct columns dynamically
        col_nome_alunos = next((col for col in df_alunos.columns if 'nome' in col.lower() and 'responsável' not in col.lower()), df_alunos.columns[1] if len(df_alunos.columns) > 1 else None)
        col_aluno_freq = next((col for col in df_freq.columns if 'aluno' in col.lower()), df_freq.columns[2] if len(df_freq.columns) > 2 else None)

        if not col_nome_alunos or not col_aluno_freq:
            raise ValueError("Não foi possível identificar as colunas de nome dos alunos nas tabelas.")

        df_alunos['__merge_key'] = df_alunos[col_nome_alunos].astype(str).str.strip().str.upper()
        df_freq['__merge_key'] = df_freq[col_aluno_freq].astype(str).str.strip().str.upper()

        merged_df = pd.merge(df_alunos, df_freq, on='__merge_key', how='inner')
        
        if merged_df.empty:
            raise ValueError("Nenhum aluno em comum foi encontrado entre os dois arquivos.")

        def find_col(df, keyword):
            return next((col for col in df.columns if keyword.lower() in col.lower()), None)

        col_data = find_col(df_freq, 'data')
        if not col_data: col_data = find_col(merged_df, 'data')
        col_turma = find_col(merged_df, 'turma')
        col_presenca = find_col(merged_df, 'presença') or find_col(merged_df, 'presenca')
        col_resp = find_col(df_alunos, 'responsável') or find_col(df_alunos, 'responsavel')
        if not col_resp: col_resp = find_col(merged_df, 'responsável')
        col_tel = find_col(merged_df, 'telefone') or find_col(merged_df, 'contato')
        col_escola = find_col(merged_df, 'escola')

        final_cols = {}
        if col_data: final_cols[col_data] = 'Data'
        if col_turma: final_cols[col_turma] = 'Turma'
        final_cols[col_nome_alunos] = 'Nome do Aluno'
        if col_presenca: final_cols[col_presenca] = 'Presença'
        if col_resp: final_cols[col_resp] = 'Nome do Responsável'
        if col_tel: final_cols[col_tel] = 'Contato'
        if col_escola: final_cols[col_escola] = 'Escola'

        final_df = merged_df[list(final_cols.keys())].rename(columns=final_cols)
        final_df = final_df.dropna(how='all')
        
        return final_df

    except Exception as e:
        raise Exception(f"Erro ao processar dados: {str(e)}")

def create_excel_buffer(df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Frequência')
        worksheet = writer.sheets['Frequência']
        
        max_row = len(df) + 1
        max_col = len(df.columns)
        
        from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
        from openpyxl.utils import get_column_letter

        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tab = ExcelTable(displayName="TabelaFrequencia", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        
        for col_num, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.column_dimensions[column_letter].width = max_length + 2

    buffer.seek(0)
    return buffer

def create_pdf_buffer(df):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    
    style_cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, textColor=colors.black)
    style_header = ParagraphStyle('Header', parent=styles['Normal'], fontSize=9, textColor=colors.whitesmoke, fontName='Helvetica-Bold')
    
    data = []
    headers = [Paragraph(escape(str(col)), style_header) for col in df.columns]
    data.append(headers)
    
    for row in df.values:
        row_data = [Paragraph(escape(str(item)) if pd.notna(item) else "", style_cell) for item in row]
        data.append(row_data)

    width, height = landscape(A4)
    usable_width = width - 40
    
    col_max_lengths = []
    for col in df.columns:
        max_len = max([len(str(item)) for item in df[col].astype(str)] + [len(str(col))])
        col_max_lengths.append(max(max_len, 5))
        
    total_len = sum(col_max_lengths)
    col_widths = [(l / total_len) * usable_width for l in col_max_lengths]

    t = Table(data, repeatRows=1, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F0F8FF")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(t)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

# Streamlit App UI
st.title("📄 Mesclador de Dados (PDF)")
st.write("Suba os arquivos PDF para mesclar as tabelas de Alunos e Frequência.")

col1, col2 = st.columns(2)

with col1:
    file1 = st.file_uploader("1. Lista de Alunos (PDF)", type=['pdf'])

with col2:
    file2 = st.file_uploader("2. Relatório de Frequência (PDF)", type=['pdf'])

if file1 and file2:
    st.info("Arquivos recebidos! Clique no botão abaixo para cruzar os dados.")
    
    if st.button("Cruzar Dados", type="primary", use_container_width=True):
        with st.spinner("Processando dados e gerando tabelas..."):
            try:
                processed_df = process_data(file1, file2)
                st.success(f"✅ Dados cruzados com sucesso! Encontrados {len(processed_df)} registros em comum.")
                
                # Show preview
                st.write("Visualização dos dados:")
                st.dataframe(processed_df, use_container_width=True)
                
                # Create buffers
                excel_buffer = create_excel_buffer(processed_df)
                pdf_buffer = create_pdf_buffer(processed_df)
                
                # Download buttons
                st.write("### Baixar Arquivos Mesclados")
                
                dl_col1, dl_col2 = st.columns(2)
                with dl_col1:
                    st.download_button(
                        label="⬇️ Exportar XLSX",
                        data=excel_buffer,
                        file_name="dados_mesclados.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with dl_col2:
                    st.download_button(
                        label="⬇️ Exportar PDF",
                        data=pdf_buffer,
                        file_name="dados_mesclados.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    
            except Exception as e:
                st.error(f"Erro: {str(e)}")
